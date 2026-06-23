#!/usr/bin/env python3
"""호암초등학교 드론마스터 월별 행정 파일 자동 생성 스크립트.

1차 구현 범위(엑셀 자동화):
- 수강료/교재비 파일의 전월 시트 복사 후 대상월 시트 생성
- 신규 신청자 추가 / 중단자 삭제
- 출석부 파일 복사 및 대상월 파일명으로 저장
- 활동비 청구서 파일 복사 및 대상월 파일명으로 저장
- 출석부/청구서 학생 수 불일치 경고

주의:
- 실제 열 매핑(학생명, 학년/반, 수강료 등)은 학교 양식마다 다를 수 있어
  --name-col, --header-row 등 CLI 옵션으로 조정 가능하게 설계함.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Sequence

from openpyxl import load_workbook


MONTH_SHEET_PATTERN = re.compile(r"^(0[1-9]|1[0-2])월$")


@dataclass
class StudentRow:
    values: list

    @property
    def name(self) -> str:
        return str(self.values[0]).strip() if self.values and self.values[0] is not None else ""


def validate_month(mm: str, field_name: str) -> str:
    if not re.fullmatch(r"(0[1-9]|1[0-2])", mm):
        raise ValueError(f"{field_name} 값 '{mm}' 이(가) 올바르지 않습니다. 01~12 두 자리로 입력하세요.")
    return mm


def ensure_sheet_name(mm: str) -> str:
    return f"{mm}월"


def read_csv_students(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = [row for row in reader if any(cell.strip() for cell in row)]
    return rows


def read_json_students(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError(f"JSON 형식 오류: {path} 는 배열(list)이어야 합니다.")

    normalized: list[list[str]] = []
    for item in data:
        if isinstance(item, dict):
            name = str(item.get("학생명", item.get("name", ""))).strip()
            grade = str(item.get("학년", item.get("grade", ""))).strip()
            class_no = str(item.get("반", item.get("class", ""))).strip()
            normalized.append([name, grade, class_no])
        elif isinstance(item, list):
            normalized.append([str(v).strip() for v in item])
        else:
            normalized.append([str(item).strip()])
    return [row for row in normalized if row and row[0]]


def read_student_rows(path: Path) -> list[list[str]]:
    if path.suffix.lower() == ".json":
        return read_json_students(path)
    return read_csv_students(path)


def list_month_sheets(workbook) -> list[str]:
    return [name for name in workbook.sheetnames if MONTH_SHEET_PATTERN.fullmatch(name)]


def find_latest_month_sheet(workbook) -> str | None:
    month_sheets = list_month_sheets(workbook)
    if not month_sheets:
        return None
    return max(month_sheets, key=lambda n: int(n[:2]))


def find_column_by_header(ws, header_row: int, candidates: Sequence[str]) -> int | None:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        text = str(value).strip() if value is not None else ""
        if text in candidates:
            return col
    return None


def collect_existing_names(ws, start_row: int, name_col: int) -> set[str]:
    names: set[str] = set()
    for r in range(start_row, ws.max_row + 1):
        value = ws.cell(row=r, column=name_col).value
        if value is None:
            continue
        name = str(value).strip()
        if name:
            names.add(name)
    return names


def remove_names(ws, start_row: int, name_col: int, names_to_remove: set[str]) -> int:
    removed = 0
    row = ws.max_row
    while row >= start_row:
        value = ws.cell(row=row, column=name_col).value
        name = str(value).strip() if value is not None else ""
        if name and name in names_to_remove:
            ws.delete_rows(row, 1)
            removed += 1
        row -= 1
    return removed


def append_students(
    ws,
    start_row: int,
    name_col: int,
    grade_col: int | None,
    class_col: int | None,
    new_students: Iterable[list[str]],
) -> int:
    existing = collect_existing_names(ws, start_row, name_col)
    added = 0
    for student in new_students:
        name = student[0].strip() if student and student[0] else ""
        if not name or name in existing:
            continue
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=name_col, value=name)
        if grade_col and len(student) > 1:
            ws.cell(row=next_row, column=grade_col, value=student[1])
        if class_col and len(student) > 2:
            ws.cell(row=next_row, column=class_col, value=student[2])
        existing.add(name)
        added += 1
    return added


def count_students(ws, start_row: int, name_col: int) -> int:
    count = 0
    for r in range(start_row, ws.max_row + 1):
        value = ws.cell(row=r, column=name_col).value
        if value is not None and str(value).strip():
            count += 1
    return count


def clone_month_sheet(
    workbook_path: Path,
    prev_mm: str,
    target_mm: str,
    header_row: int,
    data_start_row: int,
    name_col: int | None,
    grade_col: int | None,
    class_col: int | None,
    add_students: list[list[str]],
    stop_students: list[list[str]],
) -> int:
    wb = load_workbook(workbook_path)
    prev_sheet = ensure_sheet_name(prev_mm)
    target_sheet = ensure_sheet_name(target_mm)

    if prev_sheet not in wb.sheetnames:
        latest = find_latest_month_sheet(wb)
        if latest is None:
            raise ValueError(f"복사 기준 시트를 찾을 수 없습니다: {workbook_path}")
        prev_sheet = latest

    if target_sheet in wb.sheetnames:
        raise ValueError(f"대상 시트가 이미 존재합니다: {target_sheet} ({workbook_path.name})")

    source = wb[prev_sheet]
    copied = wb.copy_worksheet(source)
    copied.title = target_sheet

    inferred_name_col = name_col
    inferred_grade_col = grade_col
    inferred_class_col = class_col

    if inferred_name_col is None:
        inferred_name_col = find_column_by_header(copied, header_row, ["학생명", "성명", "이름"])
    if inferred_grade_col is None:
        inferred_grade_col = find_column_by_header(copied, header_row, ["학년"])
    if inferred_class_col is None:
        inferred_class_col = find_column_by_header(copied, header_row, ["반"])

    if inferred_name_col is None:
        raise ValueError(
            "학생명 열을 찾지 못했습니다. --name-col 값을 지정하거나 헤더명을 '학생명/성명/이름'으로 맞춰주세요."
        )

    stop_names = {row[0].strip() for row in stop_students if row and row[0].strip()}
    add_names = {row[0].strip() for row in add_students if row and row[0].strip()}
    overlaps = sorted(stop_names & add_names)
    if overlaps:
        print(f"[경고] 신규/중단 동시 존재 학생: {', '.join(overlaps)}")

    removed = remove_names(copied, data_start_row, inferred_name_col, stop_names)
    added = append_students(
        copied,
        data_start_row,
        inferred_name_col,
        inferred_grade_col,
        inferred_class_col,
        add_students,
    )

    duplicates = count_duplicates(copied, data_start_row, inferred_name_col)
    if duplicates:
        print(f"[경고] 중복 학생명 감지: {', '.join(duplicates)}")

    student_count = count_students(copied, data_start_row, inferred_name_col)

    wb.save(workbook_path)
    print(f"[완료] {workbook_path.name}: {prev_sheet} → {target_sheet}, 추가 {added}명, 삭제 {removed}명, 총 {student_count}명")
    return student_count


def count_duplicates(ws, start_row: int, name_col: int) -> list[str]:
    freq: dict[str, int] = {}
    for r in range(start_row, ws.max_row + 1):
        value = ws.cell(row=r, column=name_col).value
        name = str(value).strip() if value is not None else ""
        if not name:
            continue
        freq[name] = freq.get(name, 0) + 1
    return sorted([name for name, cnt in freq.items() if cnt > 1])


def copy_template_with_month_name(template: Path, output: Path) -> None:
    if not template.exists():
        raise FileNotFoundError(f"양식 파일이 없습니다: {template}")
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    print(f"[완료] 파일 생성: {output.name}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="호암초 월별 파일 자동 생성기 (1차: 엑셀 중심)")

    parser.add_argument("--base-path", type=Path, required=True, help="작업 기준 폴더")
    parser.add_argument("--target-month", required=True, help="대상월 (MM)")
    parser.add_argument("--prev-month", required=True, help="전월 (MM)")

    parser.add_argument(
        "--tuition-file",
        default="2026-수강료및교재(재료)비내역-드론마스터.xlsx",
        help="수강료/교재비 파일명",
    )
    parser.add_argument("--attendance-template", default="202603-출석부-드론마스터.xlsx", help="출석부 템플릿 파일명")
    parser.add_argument(
        "--invoice-template",
        default="202603 방과후학교 프로그램 활동비 청구서_드론마스터.xlsx",
        help="활동비 청구서 템플릿 파일명",
    )

    parser.add_argument("--new-students", type=Path, help="신규 신청자 CSV/JSON 경로")
    parser.add_argument("--stopped-students", type=Path, help="중단자 CSV/JSON 경로")

    parser.add_argument("--header-row", type=int, default=1, help="헤더 행 번호")
    parser.add_argument("--data-start-row", type=int, default=2, help="학생 데이터 시작 행 번호")
    parser.add_argument("--name-col", type=int, default=None, help="학생명 열 번호(1부터)")
    parser.add_argument("--grade-col", type=int, default=None, help="학년 열 번호(1부터)")
    parser.add_argument("--class-col", type=int, default=None, help="반 열 번호(1부터)")

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    target_mm = validate_month(args.target_month, "target_month")
    prev_mm = validate_month(args.prev_month, "prev_month")

    base_path: Path = args.base_path
    tuition_path = base_path / args.tuition_file
    attendance_template = base_path / args.attendance_template
    invoice_template = base_path / args.invoice_template

    new_students = read_student_rows(args.new_students) if args.new_students else []
    stopped_students = read_student_rows(args.stopped_students) if args.stopped_students else []

    student_count = clone_month_sheet(
        workbook_path=tuition_path,
        prev_mm=prev_mm,
        target_mm=target_mm,
        header_row=args.header_row,
        data_start_row=args.data_start_row,
        name_col=args.name_col,
        grade_col=args.grade_col,
        class_col=args.class_col,
        add_students=new_students,
        stop_students=stopped_students,
    )

    yyyymm = f"2026{target_mm}"
    attendance_output = base_path / f"{yyyymm}-출석부-드론마스터.xlsx"
    invoice_output = base_path / f"{yyyymm} 방과후학교 프로그램 활동비 청구서_드론마스터.xlsx"

    copy_template_with_month_name(attendance_template, attendance_output)
    copy_template_with_month_name(invoice_template, invoice_output)

    print("\n=== 생성 결과 ===")
    print(f"생성 시트명: {ensure_sheet_name(target_mm)}")
    print(f"출석부: {attendance_output.name}")
    print(f"청구서: {invoice_output.name}")
    print(f"기준 학생 수: {student_count}명")
    print("\n[안내] 출석부/청구서의 실제 인원 반영(수식/셀 매핑)은 학교 양식 열 구조에 맞춰 후속 단계에서 연결하세요.")


if __name__ == "__main__":
    main()
